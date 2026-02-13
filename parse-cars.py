import openpyxl, json

wb = openpyxl.load_workbook('public/Car Database.xlsx', data_only=True)
ws = wb['Sheet1']

vehicles = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    brand = str(row[0]).strip() if row[0] else ''
    model = str(row[1]).strip() if row[1] else ''
    year = int(row[2]) if row[2] else None
    trim_val = str(row[3]).strip() if row[3] else ''
    sqft_25 = float(row[4]) if row[4] else 0
    sqft_50 = float(row[5]) if row[5] else 0
    sqft_75 = float(row[6]) if row[6] else 0
    sqft_100 = float(row[7]) if row[7] else 0
    if brand and model:
        vehicles.append({
            'brand': brand, 'model': model, 'year': year, 'trim': trim_val,
            'sqft_25': sqft_25, 'sqft_50': sqft_50, 'sqft_75': sqft_75, 'sqft_100': sqft_100
        })

with open('car-database.json', 'w') as f:
    json.dump(vehicles, f, indent=2)

print(f'Exported {len(vehicles)} vehicles to car-database.json')

sqfts = [v['sqft_100'] for v in vehicles if v['sqft_100'] > 0]
print(f'Sqft range: {min(sqfts):.0f} - {max(sqfts):.0f}')
print(f'Average: {sum(sqfts)/len(sqfts):.0f}')

print('\n--- Reference vehicles ---')
for v in vehicles:
    name = f"{v['brand']} {v['model']} {v['year']} {v['trim']}"
    if v['brand'] == 'Honda' and 'Civic' in v['model']:
        print(f"  {name}: {v['sqft_100']} sqft")
    elif v['brand'] == 'Toyota' and 'Camry' in v['model']:
        print(f"  {name}: {v['sqft_100']} sqft")
    elif v['brand'] == 'Ford' and 'F-150' in v['model']:
        print(f"  {name}: {v['sqft_100']} sqft")
    elif v['brand'] == 'Chevrolet' and 'Tahoe' in v['model']:
        print(f"  {name}: {v['sqft_100']} sqft")
    elif v['brand'] == 'Tesla' and 'Model 3' in v['model']:
        print(f"  {name}: {v['sqft_100']} sqft")
