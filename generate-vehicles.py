import openpyxl
import json

wb = openpyxl.load_workbook('public/Car Database.xlsx', data_only=True)
ws = wb['Sheet1']

vehicles = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    brand = str(row[0]).strip() if row[0] else ''
    model = str(row[1]).strip() if row[1] else ''
    year = int(row[2]) if row[2] else None
    trim_val = str(row[3]).strip() if row[3] else ''
    sqft_100 = float(row[7]) if row[7] else 0
    if brand and model and sqft_100 > 0:
        vehicles.append({
            'make': brand,
            'model': model,
            'year': year,
            'trim': trim_val,
            'totalSqft': sqft_100
        })

# Sort by make, model, year desc
vehicles.sort(key=lambda v: (v['make'], v['model'], -(v['year'] or 0)))

# Generate TypeScript
lines = []
lines.append('// Auto-generated from Ed\'s Car Database.xlsx — 1,178 vehicles')
lines.append('// DO NOT EDIT MANUALLY — re-run generate-vehicles.py to update')
lines.append('')
lines.append('export interface VehicleEntry {')
lines.append('  make: string;')
lines.append('  model: string;')
lines.append('  year: number;')
lines.append('  trim: string;')
lines.append('  totalSqft: number;')
lines.append('}')
lines.append('')
lines.append('export const vehicleDatabase: VehicleEntry[] = [')
for v in vehicles:
    year = v['year'] if v['year'] else 0
    # Escape double quotes and backslashes in string fields
    make_esc = v['make'].replace('\\', '\\\\').replace('"', '\\"')
    model_esc = v['model'].replace('\\', '\\\\').replace('"', '\\"')
    trim_esc = v['trim'].replace('\\', '\\\\').replace('"', '\\"')
    lines.append(f'  {{ make: "{make_esc}", model: "{model_esc}", year: {year}, trim: "{trim_esc}", totalSqft: {v["totalSqft"]} }},')
lines.append('];')
lines.append('')

# Helper functions
lines.append('// Get all unique makes')
lines.append('export function getAllMakes(): string[] {')
lines.append('  const makes = new Set<string>();')
lines.append('  vehicleDatabase.forEach((v) => makes.add(v.make));')
lines.append('  return Array.from(makes).sort();')
lines.append('}')
lines.append('')
lines.append('// Get unique models for a make')
lines.append('export function getModelsForMake(make: string): string[] {')
lines.append('  const models = new Set<string>();')
lines.append('  vehicleDatabase.filter((v) => v.make === make).forEach((v) => models.add(v.model));')
lines.append('  return Array.from(models).sort();')
lines.append('}')
lines.append('')
lines.append('// Get year/trim combos for a make+model')
lines.append('export function getVariantsForModel(make: string, model: string): VehicleEntry[] {')
lines.append('  return vehicleDatabase')
lines.append('    .filter((v) => v.make === make && v.model === model)')
lines.append('    .sort((a, b) => b.year - a.year);')
lines.append('}')
lines.append('')
lines.append('// Get a specific vehicle')
lines.append('export function getVehicle(make: string, model: string, year: number, trim: string): VehicleEntry | undefined {')
lines.append('  return vehicleDatabase.find(')
lines.append('    (v) => v.make === make && v.model === model && v.year === year && v.trim === trim')
lines.append('  );')
lines.append('}')
lines.append('')

# Coverage, design tiers, finish, install, price calc
lines.append('// Coverage options')
lines.append('export const coverageOptions = [')
lines.append('  { id: "full", label: "Full Wrap", description: "Complete vehicle coverage", multiplier: 1.0 },')
lines.append('  { id: "half", label: "Half Wrap", description: "One side + hood or rear", multiplier: 0.55 },')
lines.append('  { id: "partial", label: "Partial Wrap", description: "Selected panels only", multiplier: 0.35 },')
lines.append('  { id: "hood-roof", label: "Hood + Roof", description: "Top surfaces only", multiplier: 0.2 },')
lines.append('  { id: "hood", label: "Hood Only", description: "Front hood panel", multiplier: 0.1 },')
lines.append('  { id: "rear", label: "Rear Only", description: "Trunk/tailgate area", multiplier: 0.12 },')
lines.append('  { id: "side-panels", label: "Side Panels", description: "Both door areas", multiplier: 0.4 },')
lines.append('];')
lines.append('')

lines.append('// Design tier pricing per sqft (base rate from Ed: $14/sqft)')
lines.append('export const designTiers = [')
lines.append('  {')
lines.append('    id: "premade",')
lines.append('    label: "Pre-Made Design",')
lines.append('    description: "Choose from our existing anime designs",')
lines.append('    pricePerSqft: 10,')
lines.append('    turnaround: "5-7 business days",')
lines.append('  },')
lines.append('  {')
lines.append('    id: "semicustom",')
lines.append('    label: "Semi-Custom",')
lines.append('    description: "Modify an existing design with your preferences",')
lines.append('    pricePerSqft: 14,')
lines.append('    turnaround: "10-14 business days",')
lines.append('  },')
lines.append('  {')
lines.append('    id: "fullcustom",')
lines.append('    label: "Fully Custom Itasha",')
lines.append('    description: "Original artwork designed specifically for your vehicle",')
lines.append('    pricePerSqft: 20,')
lines.append('    turnaround: "3-5 weeks",')
lines.append('  },')
lines.append('];')
lines.append('')

lines.append('// Material/finish options')
lines.append('export const finishOptions = [')
lines.append('  { id: "gloss", label: "High Gloss", priceAdd: 0 },')
lines.append('  { id: "matte", label: "Matte", priceAdd: 2 },')
lines.append('  { id: "satin", label: "Satin", priceAdd: 2 },')
lines.append('];')
lines.append('')

lines.append('// Installation option')
lines.append('export const installOptions = [')
lines.append('  { id: "ship", label: "Ship to Me (DIY)", price: 0 },')
lines.append('  { id: "install-houston", label: "Professional Install (Houston, TX)", price: 800 },')
lines.append('  { id: "install-partner", label: "Install via Partner Shop", price: 600 },')
lines.append('];')
lines.append('')

lines.append('// Calculate price')
lines.append('export function calculatePrice(')
lines.append('  sqft: number,')
lines.append('  coverageMultiplier: number,')
lines.append('  pricePerSqft: number,')
lines.append('  finishAdd: number,')
lines.append('  installPrice: number')
lines.append('): { subtotal: number; designFee: number; total: number; effectiveSqft: number } {')
lines.append('  const effectiveSqft = Math.round(sqft * coverageMultiplier);')
lines.append('  const subtotal = effectiveSqft * (pricePerSqft + finishAdd);')
lines.append('  const designFee = 99; // Refundable design fee')
lines.append('  const total = subtotal + installPrice + designFee;')
lines.append('  return { subtotal, designFee, total, effectiveSqft };')
lines.append('}')
lines.append('')

with open('src/data/vehicles.ts', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Generated src/data/vehicles.ts with {len(vehicles)} vehicles')
print(f'Makes: {len(set(v["make"] for v in vehicles))}')
print(f'Sqft range: {min(v["totalSqft"] for v in vehicles):.0f} - {max(v["totalSqft"] for v in vehicles):.0f}')
